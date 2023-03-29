from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants


class Test_OdevClass:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
        
    def teardown_method(self):
        self.driver.quit()

    def getData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sayfa1"]

        totalRows = selectedSheet.max_row
        data = []
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data

    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def test_username_required(self):
        self.waitForElementVisible((By.ID,"user-name"))
        self.waitForElementVisible((By.ID,"password"),5)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-username_required.png")
        assert errorMessage.text == "Epic sadface: Username is required"
    
    def test_password_required(self):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),5)
        usernameInput.send_keys("1")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-password_required.png")
        assert errorMessage.text == "Epic sadface: Password is required"

    def test_locked_user(self):
        self.waitForElementVisible((By.ID, "user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        passwordInput = self.driver.find_element(By.ID, "password")
        usernameInput.send_keys("locked_out_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-locked_user.png")
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."

    def test_erroriconvisible(self):
        self.waitForElementVisible((By.ID,"login-button"),5)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-error-icon-visible.png")
        elements = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(1) > .svg-inline--fa")
        assert len(elements) > 0
        elements = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(2) > .svg-inline--fa")
        assert len(elements) > 0
        elements = self.driver.find_elements(By.CSS_SELECTOR, ".fa-times > path")
        assert len(elements) > 0
        self.driver.find_element(By.CSS_SELECTOR, ".fa-times > path").click()
        self.driver.save_screenshot(f"{self.folderPath}/test-error-icon-invisible.png")
        elements = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(1) > .svg-inline--fa")
        assert len(elements) == 0

    def test_testvalidlogin(self):
        self.waitForElementVisible((By.ID, "user-name"),2)
        userInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        passwordInput = self.driver.find_element(By.ID, "password")
        userInput.send_keys("standard_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-valid_login.png")
        assert self.driver.current_url == "https://www.saucedemo.com/inventory.html"
